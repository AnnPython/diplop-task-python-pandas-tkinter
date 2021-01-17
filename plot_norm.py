import vidget_norm
import rozr_ok
from rozr_ok import *
import pandas as pd
#import xlsxwriter
from openpyxl import Workbook
from openpyxl.chart import (Reference,Series,BarChart3D,)
from openpyxl import load_workbook
#from openpyxl.styles import NamedStyle
from openpyxl.styles import Font, Color, Alignment, Border, Side, colors
from openpyxl import cell


file_name = str(period.get()) + '.xlsx'
wb = load_workbook(file_name )
sheet = wb.active
ws = wb.active

sheet.column_dimensions['A'].width = 27
sheet.column_dimensions['B'].width = 13
sheet.column_dimensions['C'].width = 13
sheet.column_dimensions['D'].width = 13
sheet.column_dimensions['E'].width = 13



thin = Side(border_style='thin', color='000000') 
black_border = Border(top=thin, left=thin, right=thin, bottom=thin)
font = Font(name='Times New Roman', size=10, bold=False, color='000000')
align = Alignment(horizontal='center', wrap_text= True, vertical='center')
font_bold = Font(name='Times New Roman', size=10, bold=True, color='000000')

for area in ['A', 'B', 'C', 'D', 'E']: 
    for col_ind in range(34):
        ind = area + str(col_ind + 1) 
        sheet[ind].alignment = align 
        sheet[ind].font = font 

for row in ws['A2:E10']:
    for cell in row:
        cell.border = black_border


for row in ws['A15:D19']:
    for cell in row:
        cell.border = black_border

for row in ws['A23:C24']:
    for cell in row:
        cell.border = black_border


for row in ws['A27:B34']:
    for cell in row:
        cell.border = black_border



sheet['A1'].font = font_bold
sheet['B1'].font = font_bold
sheet['C1'].font = font_bold
sheet['D1'].font = font_bold
sheet['E1'].font = font_bold
sheet['A14'].font = font_bold
sheet['B14'].font = font_bold
sheet['C14'].font = font_bold
sheet['D14'].font = font_bold
sheet['A26'].font = font_bold
sheet['B26'].font = font_bold
sheet['A34'].font = font_bold
sheet['B34'].font = font_bold
sheet['A22'].font = font_bold

logging.debug(f'Форматування файлу {file_name} проведено')  
   
wb.save(file_name)      


wb1 = load_workbook(file_name )
ws = wb1.active

rows = [
    ('Собівартість',vpluv_cost ),
    ('Адміністративні витрати', vpluv_admin),
    ('Інші операційні витрати', vpluv_others_expenses),
    ( 'Витрати на збут', vpluv_trade_expenses),
    ( 'Інші операційні доходи', vpluv_others_profit),
    ( 'Інші фінансові доходи', vpluv_others_finprofit),
    ( 'Дохід від реалізації' , vpluv_profit)
]

data = Reference(ws, min_col=2, min_row=26, max_col=2, max_row=33)
titles = Reference(ws, min_col=1, min_row=27, max_row=33)
chart = BarChart3D()
chart.title = 'Вплив факторів'

chart.add_data( data,titles_from_data=True)
chart.set_categories(titles)

ws.add_chart(chart, 'H4')
wb1.save(file_name)
logging.debug(f'Діаграма збережена до файлу {file_name}') 

