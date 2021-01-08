import vidget_norm
from vidget_norm import *
from openpyxl import load_workbook
import pandas as pd
import xlsxwriter
from pandas import ExcelWriter
import numpy as np



file_name = str(period.get()) + '.xlsx'
if file_name in os.listdir():
    list_dani = pd.read_excel(file_name)
    dani_priv_period = list(list_dani['Попередній період'])
    dani_now_period = list(list_dani['Поточний період'])

def read_riven():  
        
        riven_cost_price_priv=dani_priv_period[1]/(dani_priv_period[0]+dani_priv_period[6]+dani_priv_period[7])*100
        riven_cost_price_now=dani_now_period[1]/(dani_now_period[0]+dani_now_period[6]+dani_now_period[7])*100
        riven_admin_expenses_priv=dani_priv_period[3]/(dani_priv_period[0]+dani_priv_period[6]+dani_priv_period[7])*100
        riven_admin_expenses_now=dani_now_period[3]/(dani_now_period[0]+dani_now_period[6]+dani_now_period[7])*100
        riven_trade_expenses_priv=dani_priv_period[4]/(dani_priv_period[0]+dani_priv_period[6]+dani_priv_period[7])*100
        riven_trade_expenses_now=dani_now_period[4]/(dani_now_period[0]+dani_now_period[6]+dani_now_period[7])*100
        riven_others_expenses_priv=dani_priv_period[5]/(dani_priv_period[0]+dani_priv_period[6]+dani_priv_period[7])*100
        riven_others_expenses_now=dani_now_period[5]/(dani_now_period[0]+dani_now_period[6]+dani_now_period[7])*100
        riven_others_profit_priv=dani_priv_period[6]/(dani_priv_period[0]+dani_priv_period[6]+dani_priv_period[7])*100
        riven_others_profit_now=dani_priv_period[6]/(dani_now_period[0]+dani_now_period[6]+dani_now_period[7])*100

        dann={'Рівень': ['Рівень собівартості','Рівень адміністративних витрат','Рівень витрат на збут','Рівень інших операційних витрат'],
        'Попередній': [float(riven_cost_price_priv),float(riven_admin_expenses_priv), float(riven_trade_expenses_priv),float(riven_others_expenses_priv)],
        'Поточний': [float(riven_cost_price_now),float(riven_admin_expenses_now), float(riven_trade_expenses_now), float(riven_others_expenses_now)],
        'Відхилення рівня': [(float(riven_cost_price_now)-float(riven_cost_price_priv)),(float(riven_admin_expenses_now)-float(riven_admin_expenses_priv)),(float(riven_trade_expenses_now)-float(riven_trade_expenses_priv)),
                       (float(riven_others_expenses_now)-float(riven_others_expenses_priv))]}
        global df1
        df1=pd.DataFrame(dann, columns=['Рівень', 'Попередній','Поточний', 'Відхилення рівня'])
        df1=np.round(df1, 2)
        file_name = str(period.get()) + '.xlsx'
        writer = pd.ExcelWriter(file_name, engine='openpyxl')
        book = load_workbook(file_name)
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        
        df1.to_excel(writer, sheet_name='Ан', header=True, index=False,
             startcol=0,startrow=13)
        
        writer.save()
read_riven()


def koef_rentab():

    file_name = str(period.get()) + '.xlsx'
    global koef_rent_priv
    koef_rent_priv=float(dani_priv_period[8])/(float(dani_priv_period[0])+float(dani_priv_period[6])+float(dani_priv_period[7]))*100
    koef_rent_now=float(dani_now_period[8])/(float(dani_now_period[0])+float(dani_now_period[6])+float(dani_now_period[7]))*100
    rizn=koef_rent_now-koef_rent_priv
   
   
    df2=pd.DataFrame([['Коефіцієнт рентабельності'],['Попередне значення', 'Поточне значення', 'Різниця'],[float(koef_rent_priv), float(koef_rent_now), rizn]])
    df2=np.round(df2, 2)
    file_name = str(period.get()) + '.xlsx'
    writer = pd.ExcelWriter(file_name, engine='openpyxl')
    book = load_workbook(file_name)
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        
    df2.to_excel(writer, sheet_name='Ан', header=False, index=False,
             startcol=0,startrow=21)
      
    writer.save()
        
koef_rentab()    

def vpluv():  
    file_name = str(period.get()) + '.xlsx'
    if file_name in os.listdir():
        list_dani = pd.read_excel(file_name)
        dani_priv_period = list(list_dani['Попередній період'])
        dani_now_period = list(list_dani['Поточний період'])     


        global vpluv_cost
        global vpluv_admin
        global vpluv_others_expenses
        global vpluv_trade_expenses
        global vpluv_others_profit
        global vpluv_others_finprofit
        global vpluv_profit 
        
        vpluv_cost=-df1['Відхилення рівня'].iloc[0]*(dani_now_period[0]+dani_now_period[6]+dani_now_period[7])/100
                                    
        vpluv_admin=-df1['Відхилення рівня'].iloc[1]*(dani_now_period[0]+dani_now_period[6]+dani_now_period[7])/100
        vpluv_others_expenses=-df1['Відхилення рівня'].iloc[3]*(dani_now_period[0]+dani_now_period[6]+dani_now_period[7])/100
        vpluv_trade_expenses=-df1['Відхилення рівня'].iloc[2]*(dani_now_period[0]+dani_now_period[6]+dani_now_period[7])/100

        vpluv_others_profit=(dani_now_period[6]-dani_priv_period[6])*float(koef_rent_priv)/100
        vpluv_others_finprofit= (dani_now_period[7]-dani_priv_period[7])*float(koef_rent_priv)/100
        vpluv_profit=(dani_now_period[0]- dani_priv_period[0])*float(koef_rent_priv)/100
        all_vpluv=vpluv_cost+vpluv_admin+vpluv_others_expenses+vpluv_others_profit+vpluv_others_finprofit+vpluv_profit
        
        dani_vpluv={'Фактор': ['Собівартість', 'Адміністративні витрати','Інші операційні витрати','Витрати на збут',
                                                                 'Інші операційні доходи', 'Інші фінансові доходи', 'Дохід від реалізації'],
        'Вплив': [vpluv_cost, vpluv_admin,vpluv_others_expenses,vpluv_trade_expenses,vpluv_others_profit,vpluv_others_finprofit,vpluv_profit]}

        df3=pd.DataFrame(dani_vpluv, columns=['Фактор', 'Вплив'])                                    
    
        df3=np.round(df3, 2)
        file_name = str(period.get()) + '.xlsx'
        writer = pd.ExcelWriter(file_name, engine='openpyxl')
        book = load_workbook(file_name)
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        
        df3.to_excel(writer, sheet_name='Ан', header=True, index=False,
             startcol=0,startrow=25)
      
        writer.save()
vpluv()    



        


'''
dani_priv_period[0]=profit_priv
profit_now=dani_now_period[0]

others_profit_priv=dani_priv_period[6]
others_profit_now=dani_now_period[6]


others_finprofit_priv=dani_priv_period[7]
others_finprofit_now=dani_now_period[7]


cost_price_priv = dani_priv_period[1]
cost_price_now = dani_now_period[1]
admin_expenses_priv = dani_priv_period[3]
admin_expenses_now = dani_now_period[3]
trade_expenses_priv =dani_priv_period[4]
trade_expenses_now=dani_now_period[4]
others_expenses_priv=dani_priv_period[5]  
others_expenses_now= dani_now_period[5]

fin_rez_priv=dani_priv_period[8]
fin_rez_now=dani_now_period[8]
'''

